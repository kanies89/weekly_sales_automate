import sys
import os
import datetime

import numpy as np
import pandas as pd
import openpyxl

from connect import connect
from PyQt5.QtCore import pyqtSignal, QObject
from openpyxl.utils import get_column_letter


SHEETS = ['contracts_terminals', 'transactions']
TEMP = './TEMP/'
progress_count = 100


class Logger(QObject):
    # Define a custom signal for log updates
    log_updated = pyqtSignal(str)

    def __init__(self, log_file):
        super(Logger, self).__init__()
        self.terminal = sys.stdout
        self.log_file = log_file

    def write(self, message):
        self.terminal.write(message)
        self.log_file.write(message)
        self.log_updated.emit(message)

    def flush(self):
        self.terminal.flush()
        self.log_file.flush()


def week(override=None):
    if override is None:
        week_number = datetime.datetime.now().isocalendar()[1] - 1

        if week_number == 1:
            prev_week_number = 52
        else:
            prev_week_number = week_number - 1

        year = datetime.datetime.now().year

        if prev_week_number == 52:
            prev_year = datetime.datetime.now().year - 1
        else:
            prev_year = year

        print('week_number: ', week_number)
        print('prev_week_number: ', prev_week_number)
        print('year: ', year)
        print('prev_year: ', prev_year)
    else:
        print("HERE", override)
        week_number = override[0]
        year = override[1]

        if week_number == 1:
            prev_week_number = 52
            prev_year = year - 1
        else:
            prev_week_number = week_number - 1
            prev_year = datetime.datetime.now().year

    return week_number, prev_week_number, year, prev_year


def to_log():
    report_date = datetime.datetime.now().strftime("%Y-%m-%d")
    file_name = f'Log/{report_date}_LOG.txt'
    with open(file_name, 'w') as log:
        log.write(f"\nReport start --{report_date}--\n")
    return file_name


def copy_wb(from_workbook, to_workbook, dataframe, sheets, progress_callback=None):
    # Load the existing workbook
    wb = openpyxl.load_workbook(from_workbook)
    for sheet_name in sheets:
        # Set progress
        if progress_callback:
            progress_count -= 8
            step = round(100 * (1 - (progress_count / 100)), 0)
            progress_callback(int(step))

        sheet = wb[sheet_name]
        for row in dataframe[sheet_name].index:
            for col in dataframe[sheet_name].columns:
                print(col)
                coord = openpyxl.utils.get_column_letter(col + 1) + str(row + 1)
                new_value = dataframe[sheet_name].iat[row, col]

                # Handle merged cells
                for merged_range in sheet.merged_cells.ranges:
                    if coord in merged_range:
                        # Find the first cell in the merged range
                        # first_cell = merged_range.min_row, merged_range.min_col
                        # first_coord = openpyxl.utils.get_column_letter(first_cell[1]) + str(first_cell[0])
                        # wb[sheet_name][first_coord].value = new_value
                        # Merge the cells
                        # sheet.merge_cells(merged_range.coord)
                        break  # Exit the loop after setting the value and merging cells
                else:
                    # If the cell is not merged, set the value directly
                    sheet[coord].value = new_value

    # Save the changes
    wb.save(to_workbook)

    return wb


def copy_column_formatting(source_workbook, target_workbook, source_column_index, target_column_index, path, x):
    target_sheet = target_workbook[SHEETS[x]]  # Get the active sheet in the target workbook
    source_sheet = source_workbook[SHEETS[x]]
    for row in source_sheet.iter_rows(min_row=1, max_row=source_sheet.max_row):
        source_cell = source_sheet.cell(row=row[0].row, column=source_column_index)  # Specify the correct source cell
        target_cell = target_sheet.cell(row=row[0].row, column=target_column_index)  # Specify the correct target cell

        # Copy the source cell's formatting to the target cell
        target_cell.font = openpyxl.styles.Font(
            name=source_cell.font.name,
            size=source_cell.font.size,
            bold=source_cell.font.bold,
            italic=source_cell.font.italic,
            color=source_cell.font.color,
            underline=source_cell.font.underline,
            strikethrough=source_cell.font.strikethrough,
            vertAlign=source_cell.font.vertAlign,
            outline=source_cell.font.outline,
            shadow=source_cell.font.shadow,
            condense=source_cell.font.condense,
            extend=source_cell.font.extend,
        )
        target_cell.fill = openpyxl.styles.PatternFill(start_color=source_cell.fill.start_color,
                                                       end_color=source_cell.fill.end_color,
                                                       fill_type=source_cell.fill.fill_type)
        target_cell.border = openpyxl.styles.Border(
            left=source_cell.border.left,
            right=source_cell.border.right,
            top=source_cell.border.top,
            bottom=source_cell.border.bottom,
            diagonal=source_cell.border.diagonal,
            diagonal_direction=source_cell.border.diagonal_direction,
            outline=source_cell.border.outline,
            vertical=source_cell.border.vertical,
            horizontal=source_cell.border.horizontal,
        )
        target_cell.alignment = openpyxl.styles.Alignment(
            horizontal=source_cell.alignment.horizontal,
            vertical=source_cell.alignment.vertical,
            text_rotation=source_cell.alignment.text_rotation,
            wrap_text=source_cell.alignment.wrap_text,
            shrink_to_fit=source_cell.alignment.shrink_to_fit,
            indent=source_cell.alignment.indent,
        )

        # Copy cell format (number format) from source_cell to target_cell
        target_cell.number_format = source_cell.number_format

    # Save the target workbook after the loop has completed
    target_workbook.save(path)


def load_df(length, path):
    df = []
    for n in range(length):
        df.append(pd.read_csv(f'{path}{n}.csv', keep_default_na=False))
    return df


def load_or_query(length, name, temp_table, query, progress_callback=None, progress_callback_text=None):
    global progress_count
    for i in range(length):

        # Set progress
        if progress_callback:
            progress_count -= 2
            step = round(100 * (1 - (progress_count / 100)), 0)
            progress_callback(int(step))

        # Set progress text
        if progress_callback_text:
            progress_callback_text(f'Loading saved data {i + 1}/{length}.')

        if os.path.exists(f'{TEMP}{name}{i}.csv'):
            if i == (length - 1) and os.path.exists(f'{TEMP}{name}{i}.csv'):
                dataframe = load_df(length, f'{TEMP}{name}')
                break
            continue
        else:

            # Set progress text
            if progress_callback_text:
                progress_callback_text(f'Querying the Database {i + 1}/{length}.')

            dataframe = connect(temp_table, query)
            i = 0
            for df in dataframe:
                df.to_csv(f'{TEMP}{name}{i}.csv')
                i += 1
            break
    return dataframe


def create_folder_structure(address):
    # Check if the folder structure exists
    if not os.path.exists(address):
        # Create the folder structure
        os.makedirs(address)
        print("Folder structure created successfully.")
    else:
        print("Folder structure already exists.")


# Define a function to replace '-' with 0
def replace_dash_with_zero(cell_value):
    return 0 if cell_value == '-' or cell_value == '' else cell_value


# Define a function to replace '-' with 0
def replace_zero_with_dash(cell_value):
    return '-' if cell_value == 0 or cell_value == '' else cell_value


def row_sum_numeric(row):
    numeric_values = pd.to_numeric(row, errors='coerce')
    return numeric_values.sum()


def check_and_add_sum_column(date, df_workbook, column_0, column_1):
    current_year_set = True
    # Check if Sum columns exist for current year in SHEET[0]
    check_sum_col = df_workbook[SHEETS[0]].iloc[5].tolist()

    if f'{date[2]} Total' not in check_sum_col:

        current_year_set = False
        sum_column = column_0 + 1

        new_col = df_workbook[SHEETS[0]].iloc[:, sum_column - 1].copy()
        new_col.loc[5] = f'{date[2]} Total'

        old_columns = range(sum_column, df_workbook[SHEETS[0]].shape[1])
        new_columns = range(sum_column + 1, df_workbook[SHEETS[0]].shape[1] + 1)
        df_workbook[SHEETS[0]].rename(columns=dict(zip(old_columns, new_columns)), inplace=True)
        df_workbook[SHEETS[0]].insert(loc=sum_column, column=sum_column, value=new_col)

    # Check if Sum columns exist for current year in SHEET [1]
    check_sum_col = df_workbook[SHEETS[1]].iloc[5].tolist()
    if f'{date[2]} Total' not in check_sum_col:

        current_year_set = False
        sum_column = column_1 + 1

        new_col = df_workbook[SHEETS[1]].iloc[:, sum_column].copy()
        new_col.loc[5] = f'{date[2]} Total'

        old_columns = range(sum_column, df_workbook[SHEETS[1]].shape[1])
        new_columns = range(sum_column + 1, df_workbook[SHEETS[1]].shape[1] + 1)
        df_workbook[SHEETS[1]].rename(columns=dict(zip(old_columns, new_columns)), inplace=True)
        df_workbook[SHEETS[1]].insert(loc=sum_column, column=sum_column, value=new_col)

        position = sum_column + date[2] - 2019

        new_col = df_workbook[SHEETS[1]].iloc[:, position].copy()
        new_col.loc[5] = f'{date[2]} Total'

        old_columns = range(position, df_workbook[SHEETS[1]].shape[1])
        new_columns = range(position + 1, df_workbook[SHEETS[1]].shape[1] + 1)
        df_workbook[SHEETS[1]].rename(columns=dict(zip(old_columns, new_columns)), inplace=True)
        df_workbook[SHEETS[1]].insert(loc=position, column=position, value=new_col)

    return current_year_set


def to_float(value):
    # Convert the value to a float using pd.to_numeric()
    try:
        # Convert the value to a float using pd.to_numeric()
        value_float = pd.to_numeric(value, errors='coerce')  # Convert to float and replace non-numeric values with NaN

        # Replace NaN with 0
        if pd.isna(value_float):
            value_float = 0

        return value_float

    except ValueError:
        return 0  # Handle the case where the conversion fails


def automate_report(progress_callback=None, progress_callback_text=None):
    global progress_count

    create_folder_structure('./Log')
    create_folder_structure('./TEMP')

    # Set progress
    if progress_callback:
        progress_count -= 1
        step = round(100*(1-(progress_count/100)), 0)
        progress_callback(int(step))

    date = week()

    # Set progress text
    if progress_callback_text:
        progress_callback_text(f'Reading the data ...sales channel_with_txns_v3_w{date[1]}.xlsx')

    # Last report
    path_prev = f'./{date[3]}/w{date[1]}/PayTel - weekly report - Sales KPIs by sales channel_with_txns_v3_w{date[1]}.xlsx'
    # New report
    create_folder_structure(f'./{date[2]}/w{date[0]}')

    # Set progress text
    if progress_callback_text:
        progress_callback_text(f'Creating excel sheet for new data ...sales channel_with_txns_v3_w{date[0]}.xlsx')

    path_new = f'./{date[2]}/w{date[0]}/PayTel - weekly report - Sales KPIs by sales channel_with_txns_v3_w{date[0]}.xlsx'
    df_workbook = pd.read_excel(path_prev, sheet_name=SHEETS, header=None, keep_default_na=False)

    # SHEET 1
    column_0 = pd.Index(df_workbook[SHEETS[0]].iloc[5]).get_loc(f'{date[1]}_{date[3]}') + 1
    new_col = df_workbook[SHEETS[0]].iloc[:, column_0 - 1].copy()

    # Change the date in header
    new_col.loc[5] = f'{date[0]}_{date[2]}'
    print(new_col)
    increment = 1
    old_columns = range(column_0, df_workbook[SHEETS[0]].shape[1])
    new_columns = range(column_0 + increment, df_workbook[SHEETS[0]].shape[1] + increment)
    df_workbook[SHEETS[0]].rename(columns=dict(zip(old_columns, new_columns)), inplace=True)
    df_workbook[SHEETS[0]].insert(loc=column_0, column=column_0, value=new_col)
    print(date)
    # SHEET 2
    column_1 = pd.Index(df_workbook[SHEETS[1]].iloc[5]).get_loc(f'{date[1]}_{date[3]}') + 1
    new_col = df_workbook[SHEETS[1]].iloc[:, column_1 - 1].copy()
    print(get_column_letter(column_0), get_column_letter(column_1))

    # Change the date in header
    new_col.loc[5] = f'{date[0]}_{date[2]}'
    print(new_col)
    increment = 1
    old_columns = range(column_1, df_workbook[SHEETS[1]].shape[1])
    new_columns = range(column_1 + increment, df_workbook[SHEETS[1]].shape[1] + increment)
    df_workbook[SHEETS[1]].rename(columns=dict(zip(old_columns, new_columns)), inplace=True)
    df_workbook[SHEETS[1]].insert(loc=column_1, column=column_1, value=new_col)

    # Set progress text
    if progress_callback_text:
        progress_callback_text(f'Inserting new columns ...sales channel_with_txns_v3_w{date[0]}.xlsx')

    # Set progress
    if progress_callback:
        progress_count -= 1
        step = round(100*(1-(progress_count/100)), 0)
        progress_callback(int(step))

    # Check if there are sum columns for current year
    current_year = check_and_add_sum_column(date, df_workbook, column_0, column_1)

    temp_table = 'Query/Temp.sql'
    query = 'Query/Query.sql'

    # Set progress text
    if progress_callback_text:
        progress_callback_text(f'Querying the Database.')

    data = load_or_query(19, f'df_workbook_{date[1]}', temp_table, query, progress_callback, progress_callback_text)

    df_title = []
    df_views = []

    # Set progress text
    if progress_callback_text:
        progress_callback_text(f'Making the dataframe views.')

    # SHEET 0 i in range(3)
    for i in range(3):
        df_title.append(df_workbook[SHEETS[0]].iat[5 + i * 25 - 3, 1])

        # Make the view of single table
        df_view = df_workbook[SHEETS[0]][6 + i * 25:25 + i * 25]
        df_views.append(df_view)

    # SHEET 1 i in range(10)
    for i in range(10):
        df_title.append(df_workbook[SHEETS[1]].iat[5 + i * 25 - 3, 1])

        # Make the view of single table
        df_view = df_workbook[SHEETS[1]][6 + i * 25:25 + i * 25]
        df_views.append(df_view)

    # SHEET 1 i in range(5)
    for i in range(5):
        df_title.append(df_workbook[SHEETS[1]].iat[255 + i * 24 - 3, 1])

        # Make the view of single table
        df_view = df_workbook[SHEETS[1]][256 + i * 24:274 + i * 24]
        df_views.append(df_view)

    # SHEET 1 i in range(2)
    for i in range(2):
        df_title.append(df_workbook[SHEETS[1]].iat[377 + i * 26 - 3, 1])

        # Make the view of single table
        df_view = df_workbook[SHEETS[1]][378 + i * 26:396 + i * 26]
        df_views.append(df_view)

    # SHEET 1 i in range(3)
    for i in range(3):
        df_title.append(df_workbook[SHEETS[1]].iat[427 + i * 24 - 3, 1])

        # Make the view of single table
        df_view = df_workbook[SHEETS[1]][428 + i * 24:446 + i * 24]
        df_views.append(df_view)

    k = 0
    # Add new values taken from Query
    for v, df_view in enumerate(df_views):

        # Set progress text
        if progress_callback_text:
            progress_callback_text(f'Adding data to new tables (table {v + 1} / {len(df_views)}')

        if v not in (13, 15, 16, 17, 18, 20, 21, 22):
            for j in range(df_view.shape[0] - 1):
                df_view.iat[j, column_0] = round(to_float(data[k].iat[j, data[k].shape[1] - 1]), 0)
            k += 1
        elif v == 13:
            for j in range(df_view.shape[0]):
                df_view.iat[j, column_0] = round(to_float(data[k].iat[j, data[k].shape[1] - 1]), 0)

            k += 1
        elif v == 20 or v == 17 or v == 18:
            for j in range(df_view.shape[0] - 1):
                value_to_convert = data[k].iat[j, data[k].shape[1] - 1]
                if value_to_convert:
                    try:
                        df_view.iat[j, column_0] = float(value_to_convert)
                    except ValueError:
                        # Handle the case where conversion fails
                        df_view.iat[j, column_0] = '-'
                elif pd.isna(value_to_convert):
                    df_view.iat[j, column_0] = '-'
            k += 1
        elif v == 15:
            for j in range(df_view.shape[0]):
                if pd.isna(df_view.iat[j, column_0]) or df_view.iat[j, column_0] == 0:
                    df_view.iat[j, column_0] = 0
                else:
                    try:
                        df_view.iat[j, column_0] = to_float(df_views[5].iat[j, column_0 - 1]) / to_float(df_views[3].iat[j, column_0])
                    except TypeError:
                        df_view.iat[j, column_0] = 0
        elif v == 16:
            for j in range(df_view.shape[0]):
                if pd.isna(df_views[6].iat[j, column_0]) or df_views[6].iat[j, column_0] == 0:
                    df_view.iat[j, column_0] = 0
                else:
                    df_view.iat[j, column_0] = df_views[6].iat[j, column_0] / df_views[4].iat[j, column_0]
        elif v == 21:
            for j in range(df_view.shape[0]):
                if pd.isna(df_views[1].iat[j, column_0]) is np.NAN or df_views[1].iat[j, column_0] == 0:
                    df_view.iat[j, column_0] = '-'
                else:
                    df_view.iat[j, column_0] = df_views[19].iat[j, column_0] / df_views[1].iat[j, column_0]
        elif v == 22:
            for j in range(df_view.shape[0]):
                if pd.isna(df_views[2].iat[j, column_0]) is np.NAN or df_views[2].iat[j, column_0] == 0:
                    df_view.iat[j, column_0] = '-'
                else:
                    df_view.iat[j, column_0] = df_views[20].iat[j, column_0] / df_views[2].iat[j, column_0]
        else:
            print('Error')

    # Set progress
    if progress_callback:
        progress_count -= 5
        step = round(100*(1-(progress_count/100)), 0)
        progress_callback(int(step))

    # Sum values.
    # Add new values using formula
    for v, df_view in enumerate(df_views):

        # Set progress text
        if progress_callback_text:
            progress_callback_text(f'Adding Total Sum Values (table number {v})')

        if v not in (7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 21, 22):
            df_sum = df_view.iloc[:, column_0 + 1 - date[0]: column_0 + 1]
            df_sum_sum = df_sum.apply(row_sum_numeric, axis=1)

            for j in range(df_view.shape[0]):
                df_view.iat[j, column_0 + 1] = df_sum_sum.iloc[j]

        elif v in (13, 14, 15, 16, 17, 18):
            for j in range(df_view.shape[0]):
                df_numeric = pd.to_numeric(df_view.iloc[j, column_0 + 1 - date[0]: column_0 + 1], errors='coerce')
                df_view.iat[j, column_0 + 1] = df_numeric.mean(numeric_only=True)

        elif v == 21:
            if pd.isna(df_views[1].iat[j, column_0 + 1]) is np.NAN or df_views[1].iat[j, column_0 + 1] == 0:
                df_view.iat[j, column_0] = '-'
            else:
                for j in range(df_view.shape[0]):
                    df_view.iat[j, column_0] = df_views[19].iat[j, column_0 + 1] / df_views[1].iat[j, column_0 + 1]

        elif v == 22:
            if pd.isna(df_views[2].iat[j, column_0] + 1) is np.NAN or df_views[2].iat[j, column_0 + 1] == 0:
                df_view.iat[j, column_0] = '-'
            else:
                for j in range(df_view.shape[0]):
                    df_view.iat[j, column_0] = df_views[20].iat[j, column_0 + 1] / df_views[2].iat[j, column_0 + 1]

    # Sum percentage values.
    # Add new values using formula
    for v, df_view in enumerate(df_views):
        if v in (3, 4, 5, 6):

            # Set progress text
            if progress_callback_text:
                progress_callback_text(f'Adding Total Sum Percentage Values (table number {v})')

            df_sum = df_view.iloc[:df_view.shape[0] - 1, column_0 + 1]
            df_sum_sum = df_sum.sum()

            for j in range(df_view.shape[0] - 1):
                df_view.iat[j, column_0 + 1 + (date[2] - 2019)] = df_view.iat[j, column_0 + 1] / df_sum_sum

    # Bottom sum.
    # Add new values using formula
    for v, df_view in enumerate(df_views):

        # Set progress text
        if progress_callback_text:
            progress_callback_text(f'Adding Bottom Sum (table {v})')

        if v in range(3):
            sheet = 0
        elif v in range(3, 7):
            sheet = 1
        if v in range(7):
            df_sum_1 = df_view.iloc[:df_view.shape[0] - 1, column_0]
            df_sum_2 = df_view.iloc[:df_view.shape[0] - 1, column_0 + 1]
            index = df_view.iloc[:df_view.shape[0] - 1, column_0].index[-1] + 1

            df_sum_sum_1 = []
            df_sum_sum_2 = []

            for i in df_sum_1:
                if not isinstance(i, str):
                    df_sum_sum_1.append(i)
            for i in df_sum_2:
                if not isinstance(i, str):
                    df_sum_sum_2.append(i)

            total_sum_1 = sum(df_sum_sum_1)
            total_sum_2 = sum(df_sum_sum_2)

            df_workbook[SHEETS[sheet]].iat[index, column_0] = total_sum_1
            df_workbook[SHEETS[sheet]].iat[index, column_0 + 1] = total_sum_2

        if v in range(7, 13):
            df_sum_1 = df_view.iloc[:df_view.shape[0] - 1, column_0]

            index = df_view.iloc[:df_view.shape[0] - 1, column_0].index[-1] + 1

            df_sum_sum_1 = []

            for i in df_sum_1:
                if not isinstance(i, str):
                    df_sum_sum_1.append(i)

            total_sum_1 = sum(df_sum_sum_1)

            df_workbook[SHEETS[sheet]].iat[index, column_0] = total_sum_1

    for v, df_view in enumerate(df_views):
        index = df_view.iloc[:df_view.shape[0] - 1, column_0].index[0] - 1
        df_workbook[SHEETS[sheet]].iat[index, column_0] = f'{date[0]}_{date[2]}'

        if date[0] == 111212121:
            print("HERE")
            col_name = f'{date[3]} Total'
            try:
                col_index = pd.Index(df_workbook[SHEETS[0]].iloc[5]).get_loc(col_name)
            except KeyError:
                print("Not found")

            df_workbook[SHEETS[0]].insert(col_index, 'New_Empty_Column', value=pd.NA)
            df_workbook[SHEETS[0]].iat[5, col_index] = f' {date[2]} Total'
            print("HERE")

        # If current year False -> no  "current year Total" column, then column was added and needs to change the header
        if current_year != True:
            print('??here??')
            col_name = f'{date[2]} Total'
            column = pd.Index(df_workbook[SHEETS[0]].iloc[5]).get_loc(col_name)
            index = df_view.iloc[:df_view.shape[0] - 1, column].index[0] - 1
            df_workbook[SHEETS[0]].iat[index, column - 1] = col_name

            if v in range(3, 7):
                col_name = f'{date[2]}'
                column = pd.Index(df_workbook[SHEETS[1]].iloc[5]).get_loc(col_name) + date[2] - 2019
                index = df_view.iloc[:df_view.shape[0] - 1, column].index[0] - 1
                df_workbook[SHEETS[1]].iat[index, column - 1] = col_name
            else:
                break

    # Set progress
    if progress_callback:
        progress_count -= 5
        step = round(100*(1-(progress_count/100)), 0)
        progress_callback(int(step))

    df_workbook[SHEETS[0]].to_excel('1.xlsx')
    df_workbook[SHEETS[1]].to_excel('2.xlsx')

    # Set progress text
    if progress_callback_text:
        progress_callback_text(f'Adding saving copy.')

    wb = copy_wb(path_prev, path_new, df_workbook, SHEETS)

    target_workbook = openpyxl.load_workbook(path_new)
    source_workbook = openpyxl.load_workbook(path_prev)

    # Set progress text
    if progress_callback_text:
        progress_callback_text(f'Copying table format for SHEET 0')

    # SHEET[0]
    for i in range(1, column_0 + 1):
        copy_column_formatting(source_workbook, target_workbook, i, i, path_new, 0)
    # Set progress
    if progress_callback:
        progress_count -= 14
        step = round(100*(1-(progress_count/100)), 0)
        progress_callback(int(step))

    # Copy formatting from column H (188) to column I (189) for the target
    copy_column_formatting(source_workbook, target_workbook, column_0, column_0 + 1, path_new, 0)

    # Copy formatting from column I (189) to the last column (190 to the end) for the target
    last_column = wb[SHEETS[0]].max_column

    for i in range(column_0, last_column):
        print(1 + i)
        copy_column_formatting(source_workbook, target_workbook, i, 1 + i, path_new, 0)
    # Set progress
    if progress_callback:
        progress_count -= 15
        step = round(100*(1-(progress_count/100)), 0)
        progress_callback(int(step))

    # Set progress text
    if progress_callback_text:
        progress_callback_text(f'Copying table format for SHEET 1')

    # SHEET[1]
    for i in range(1, column_1 + 1):
        copy_column_formatting(source_workbook, target_workbook, i, i, path_new, 1)

    # Set progress
    if progress_callback:
        progress_count -= 8
        step = round(100*(1-(progress_count/100)), 0)
        progress_callback(int(step))

    # Copy formatting from column H (188) to column I (189) for the target
    copy_column_formatting(source_workbook, target_workbook, column_1, column_1 + 1, path_new, 1)

    # Copy formatting from column I (189) to the last column (190 to the end) for the target
    last_column = 2 * (datetime.datetime.now().year - 2019) + 1
    for i in range(0, last_column):
        copy_column_formatting(source_workbook, target_workbook, column_1 + i, column_1 + 1 + i, path_new, 1)

    # Set progress text
    if progress_callback_text:
        progress_callback_text(f'Saving...')

    # Save the updated workbook
    target_workbook.save(path_new)
    # Set progress
    if progress_callback:
        progress_callback(100)

    # Set progress text
    if progress_callback_text:
        progress_callback_text(f'Finished.')


if __name__ == '__main__':
    automate_report(progress_callback=None)
